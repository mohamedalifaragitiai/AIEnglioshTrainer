"""Report generation — JSON, CSV, Excel (xlsx), PDF.

Assembles a learner's headline profile, per-dimension assessment history, ranked
gaps, adaptive plan, and feedback into one report, then renders it to the requested
format (bytes). Offline/pure-Python: openpyxl for xlsx, reportlab for PDF. Rendered
files are also written under the report dir and recorded in the reports table.
"""

from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass

from backend.coldpath.scoring import DIMENSIONS
from backend.domain.models import (
    Assessment,
    Feedback,
    GapItem,
    Plan,
    ProgressOverview,
)

REPORT_FORMATS = ("json", "csv", "xlsx", "pdf")

_CONTENT_TYPES = {
    "json": "application/json",
    "csv": "text/csv",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}


def content_type(fmt: str) -> str:
    return _CONTENT_TYPES[fmt]


@dataclass
class ReportData:
    overview: ProgressOverview
    assessments: list[Assessment]
    gaps: list[GapItem]
    plan: Plan
    feedback: Feedback


def _payload(data: ReportData) -> dict:
    return {
        "overview": data.overview.model_dump(),
        "gaps": [g.model_dump() for g in data.gaps],
        "plan": data.plan.model_dump(),
        "feedback": data.feedback.model_dump(),
        "assessments": [a.model_dump() for a in data.assessments],
    }


def render_json(data: ReportData) -> bytes:
    return json.dumps(_payload(data), indent=2).encode("utf-8")


def render_csv(data: ReportData) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["created_at", "overall", *DIMENSIONS, "scoring_model_version"])
    for a in data.assessments:
        writer.writerow(
            [a.created_at, a.overall, *[getattr(a, d) for d in DIMENSIONS], a.scoring_model_version]
        )
    return buf.getvalue().encode("utf-8")


def render_xlsx(data: ReportData) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ov = data.overview

    ws = wb.active
    ws.title = "Summary"
    ws.append(["Learner", ov.display_name, f"({ov.user_id})"])
    ws.append(["Level", ov.current_level, f"next: {ov.next_level}"])
    ws.append(["Streak (days)", ov.streak_days])
    ws.append(["Latest overall", ov.latest_overall])
    ws.append(["Est. days to next level", ov.estimated_days_to_next_level])
    ws.append([])
    ws.append(["Plan"])
    ws.append([data.plan.summary])
    for fa in data.plan.focus_areas:
        ws.append([fa.skill, fa.score, fa.why])

    wa = wb.create_sheet("Assessments")
    wa.append(["created_at", "overall", *DIMENSIONS])
    for a in data.assessments:
        wa.append([a.created_at, a.overall, *[getattr(a, d) for d in DIMENSIONS]])

    wg = wb.create_sheet("Gaps")
    wg.append(["rank", "skill", "score", "target", "gap", "severity"])
    for g in data.gaps:
        wg.append([g.rank, g.skill, g.score, g.target, g.gap, g.severity])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _fit(s: str, font: str, size: float, avail: float) -> str:
    """Trim to what actually fits, measured in points.

    The old cap was a flat 118 characters, which is width-blind in a proportional
    font: a long line ran past the right edge and appeared cut mid-word instead of
    ending cleanly.
    """
    from reportlab.pdfbase.pdfmetrics import stringWidth

    if stringWidth(s, font, size) <= avail:
        return s
    while s and stringWidth(s + "…", font, size) > avail:
        s = s[:-1]
    return s + "…"


def _wrap(s: str, font: str, size: float, avail: float, max_lines: int) -> list[str]:
    """Greedy word wrap, capped at max_lines with the last line ellipsized."""
    from reportlab.pdfbase.pdfmetrics import stringWidth

    lines: list[str] = []
    line = ""
    for word in str(s).split():
        cand = f"{line} {word}".strip()
        if not line or stringWidth(cand, font, size) <= avail:
            line = cand
            continue
        lines.append(line)
        line = word
        if len(lines) == max_lines:
            break
    if line and len(lines) < max_lines:
        lines.append(line)
    if not lines:
        return [""]
    lines[-1] = _fit(lines[-1], font, size, avail)
    return lines


def next_level_line(next_level: int | None, eta_days: int | None) -> str:
    """Header phrasing for a learner's next level.

    Three cases, not two. Requiring BOTH values before saying anything meant a
    learner on level 1 of 5 with too little trend data to estimate a date was told
    "At the top level" — the ETA being unknown is not the same as there being no
    next level to reach.
    """
    if next_level is None:
        return "At the top level"
    if eta_days is not None:
        return f"Next: level {next_level} in ~{eta_days} days"
    return f"Next: level {next_level}"


def render_pdf(data: ReportData) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas

    from backend.coldpath.scoring import level_name

    ov, fb, plan = data.overview, data.feedback, data.plan
    latest = data.assessments[-1] if data.assessments else None
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    W, H = A4
    M = 2 * cm
    y = H - 1.6 * cm
    teal, ink, grey = (0.09, 0.55, 0.5), (0.12, 0.16, 0.24), (0.42, 0.47, 0.55)

    def text(s, size=10.5, dy=0.52 * cm, bold=False, color=ink, indent=0.0):
        nonlocal y
        font = "Helvetica-Bold" if bold else "Helvetica"
        c.setFillColorRGB(*color)
        c.setFont(font, size)
        c.drawString(M + indent, y, _fit(str(s), font, size, W - 2 * M - indent))
        y -= dy

    def heading(s):
        nonlocal y
        y -= 0.15 * cm
        c.setFillColorRGB(*teal)
        c.rect(M, y - 0.05 * cm, 0.28 * cm, 0.42 * cm, fill=1, stroke=0)
        c.setFillColorRGB(*ink)
        c.setFont("Helvetica-Bold", 12.5)
        c.drawString(M + 0.45 * cm, y, s)
        y -= 0.62 * cm

    def bar(skill, pct):
        nonlocal y
        c.setFillColorRGB(*ink)
        c.setFont("Helvetica", 10)
        c.drawString(M, y, skill.capitalize())
        c.setFillColorRGB(*grey)
        c.drawRightString(W - M, y, f"{pct:.0f}%")
        bx, bw = M + 3.2 * cm, W - 2 * M - 4.6 * cm
        c.setFillColorRGB(0.86, 0.89, 0.94)
        c.roundRect(bx, y - 0.05 * cm, bw, 0.28 * cm, 0.14 * cm, fill=1, stroke=0)
        c.setFillColorRGB(*teal)
        c.roundRect(bx, y - 0.05 * cm, max(0.14 * cm, bw * pct / 100), 0.28 * cm,
                    0.14 * cm, fill=1, stroke=0)
        y -= 0.56 * cm

    # header band
    c.setFillColorRGB(*teal)
    c.rect(0, H - 1.15 * cm, W, 1.15 * cm, fill=1, stroke=0)
    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica-Bold", 15)
    c.drawString(M, H - 0.78 * cm, "AI English Coach — Progress Report")
    y = H - 1.9 * cm

    text(f"{ov.display_name}", 15, 0.5 * cm, bold=True)
    text(f"Level {ov.current_level}/5 — {level_name(ov.current_level)}", 11.5, 0.5 * cm, color=teal)
    overall = round(ov.latest_overall) if ov.latest_overall is not None else "-"
    nxt = next_level_line(ov.next_level, ov.estimated_days_to_next_level)
    text(f"Overall {overall}%   ·   Streak {ov.streak_days} days   ·   {ov.assessments_count} "
         f"assessments   ·   {nxt}", 10, 0.7 * cm, color=grey)

    heading("Skill breakdown")
    if latest:
        for d in DIMENSIONS:
            v = getattr(latest, d)
            if v is not None:
                bar(d, float(v))
    else:
        text("No assessments yet.", color=grey)

    heading("What you're doing well")
    strengths = ", ".join(s.capitalize() for s in fb.strengths) or "Keep practicing to build these."
    text("· " + strengths)
    heading("Focus areas")
    text("· " + (", ".join(w.capitalize() for w in fb.weaknesses) or "All skills near target."))
    for fa in plan.focus_areas:
        act = fa.activities[0] if fa.activities else ""
        text(f"  – {fa.skill.capitalize()} ({fa.score:.0f}%): {act}", 9.5, 0.48 * cm, color=grey)

    if fb.corrections:
        heading("Corrections from your latest session")
        for corr in fb.corrections[:5]:
            t, fix = corr.get("text", ""), corr.get("correction", "")
            # Plain ASCII arrow: the ✗ / → / ✓ glyphs are not in Helvetica's WinAnsi
            # encoding, so they rendered as nothing and left the two phrases running
            # together with only whitespace between them.
            line = f"{t}  ->  {fix}" if fix else str(t)
            for ln in _wrap(line, "Helvetica", 9.5, W - 2 * M - 0.4 * cm, 2):
                text(ln, 9.5, 0.44 * cm, indent=0.4 * cm)
    if fb.vocabulary_suggestions:
        heading("Vocabulary to try")
        text("· " + ", ".join(fb.vocabulary_suggestions[:8]), 10, color=grey)
    if fb.pronunciation_tip:
        heading("Pronunciation tip")
        text("· " + fb.pronunciation_tip, 10, color=grey)

    heading("Your plan")
    text(plan.summary, 10, 0.5 * cm)
    horizon = plan.horizon.replace("_", " ")
    text(f"Difficulty {plan.difficulty} · horizon {horizon}", 9.5, color=grey)

    c.setFillColorRGB(*grey)
    c.setFont("Helvetica-Oblique", 8.5)
    ver = latest.scoring_model_version if latest else "v1"
    c.drawString(M, 1.2 * cm, f"Generated by AI English Coach · scoring {ver}")
    c.showPage()
    c.save()
    return buf.getvalue()


def render(data: ReportData, fmt: str) -> bytes:
    if fmt == "json":
        return render_json(data)
    if fmt == "csv":
        return render_csv(data)
    if fmt == "xlsx":
        return render_xlsx(data)
    if fmt == "pdf":
        return render_pdf(data)
    raise ValueError(f"unknown report format: {fmt!r}")
